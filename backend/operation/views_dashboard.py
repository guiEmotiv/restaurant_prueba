from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, Avg, Prefetch
from django.http import JsonResponse
from datetime import datetime, date
from decimal import Decimal
from .models import Order, OrderItem, Payment, ContainerSale
from inventory.models import Recipe, RecipeItem, Ingredient

class DashboardViewSet(viewsets.ViewSet):
    """
    ViewSet para el dashboard consolidado con datos detallados completos
    Cada fila representa un item de orden con TODOS sus detalles e ingredientes
    """
    permission_classes = [AllowAny]  # Acceso completo en desarrollo
    
    @action(detail=False, methods=['get'])
    def report(self, request):
        """
        Endpoint consolidado que devuelve datos detallados por fila con filtro inteligente de período
        Cada fila incluye: Orden, Item, Receta, Ingredientes, Costos, Mesa, Zona, Pago, etc.
        """
        try:
            # Obtener parámetros de período inteligente - compatible con Django request y DRF request
            query_params = getattr(request, 'query_params', request.GET)
            period = query_params.get('period', 'all')  # Si no se especifica período, mostrar todos los datos
            date_param = query_params.get('date')
            
            # Calcular fechas según el período
            period_dates = self._calculate_period_dates(period, date_param)
            
            # Obtener datos detallados completos filtrados por período
            detailed_data = self._get_detailed_order_data_period(period_dates)
            
            # Obtener TODOS los datos para la gráfica de estados
            all_detailed_data = self._get_all_detailed_order_data_period(period_dates)
            
            # Generar dashboard desde los datos detallados
            dashboard_data = self._generate_dashboard_from_details(detailed_data, period_dates['display_date'])
            
            # Agregar información del período
            dashboard_data['period_info'] = {
                'period': period,
                'start_date': period_dates['start_date'].isoformat() if period_dates['start_date'] else None,
                'end_date': period_dates['end_date'].isoformat() if period_dates['end_date'] else None,
                'display_date': period_dates['display_date'].isoformat(),
                'total_days': period_dates['total_days']
            }
            
            # Agregar distribución de estados de items
            dashboard_data['item_status_breakdown'] = {
                'CREATED': len([row for row in all_detailed_data if row['item_status'] == 'CREATED']),
                'PREPARING': len([row for row in all_detailed_data if row['item_status'] == 'PREPARING']),
                'SERVED': len([row for row in all_detailed_data if row['item_status'] == 'SERVED']),
                'PAID': len([row for row in all_detailed_data if row['item_status'] == 'PAID'])
            }
            
            # Agregar recetas no vendidas para el período
            dashboard_data['unsold_recipes'] = self._get_unsold_recipes_period(detailed_data, period_dates)
            
            return Response(dashboard_data)
        
        except Exception as e:
            return Response({
                'error': f'Error processing dashboard request: {str(e)}',
                'date': None,
                'detailed_data': [],
                'summary': {'total_orders': 0, 'total_revenue': 0, 'average_ticket': 0, 'average_service_time': 0},
                'category_breakdown': [], 'top_dishes': [], 'waiter_performance': [],
                'zone_performance': [], 'top_tables': [], 'payment_methods': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _calculate_period_dates(self, period, date_param=None):
        """
        Calcula las fechas de inicio y fin según el período seleccionado
        """
        from datetime import timedelta
        
        today = timezone.now().date()
        
        if date_param:
            try:
                reference_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                reference_date = today
        else:
            reference_date = today
        
        if period == 'today':
            return {
                'start_date': reference_date,
                'end_date': reference_date,
                'display_date': reference_date,
                'total_days': 1
            }
        elif period == 'yesterday':
            yesterday = reference_date - timedelta(days=1)
            return {
                'start_date': yesterday,
                'end_date': yesterday,
                'display_date': yesterday,
                'total_days': 1
            }
        elif period == 'week':
            # Última semana (7 días)
            start_date = reference_date - timedelta(days=6)
            return {
                'start_date': start_date,
                'end_date': reference_date,
                'display_date': reference_date,
                'total_days': 7
            }
        elif period == 'month':
            # Último mes (30 días)
            start_date = reference_date - timedelta(days=29)
            return {
                'start_date': start_date,
                'end_date': reference_date,
                'display_date': reference_date,
                'total_days': 30
            }
        elif period == 'all':
            # Todos los datos
            return {
                'start_date': None,
                'end_date': None,
                'display_date': reference_date,
                'total_days': None
            }
        else:
            # Por defecto: hoy
            return {
                'start_date': reference_date,
                'end_date': reference_date,
                'display_date': reference_date,
                'total_days': 1
            }
    
    def _get_detailed_order_data_period(self, period_dates):
        """
        Obtiene datos detallados filtrados por período
        """
        # Base query para órdenes PAID
        paid_orders_query = Order.objects.filter(status='PAID')
        
        # Aplicar filtro de fechas si no es 'all'
        if period_dates['start_date'] and period_dates['end_date']:
            paid_orders_query = paid_orders_query.filter(
                paid_at__date__gte=period_dates['start_date'],
                paid_at__date__lte=period_dates['end_date']
            )
        
        paid_orders = paid_orders_query.select_related(
            'table__zone'
        ).prefetch_related(
            Prefetch('orderitem_set', queryset=OrderItem.objects.filter(
                status='PAID'  # SOLO ITEMS PAGADOS
            ).select_related(
                'recipe__group'
            ).prefetch_related(
                'recipe__recipeitem_set__ingredient__unit'
            )),
            'payments',
            'container_sales__container'
        ).order_by('paid_at')
        
        return self._process_detailed_order_data(paid_orders)
    
    def _get_all_detailed_order_data_period(self, period_dates):
        """
        Obtiene TODOS los datos detallados (incluyendo items no PAID) filtrados por período
        """
        # Base query para órdenes PAID
        paid_orders_query = Order.objects.filter(status='PAID')
        
        # Aplicar filtro de fechas si no es 'all'
        if period_dates['start_date'] and period_dates['end_date']:
            paid_orders_query = paid_orders_query.filter(
                paid_at__date__gte=period_dates['start_date'],
                paid_at__date__lte=period_dates['end_date']
            )
        
        paid_orders = paid_orders_query.select_related(
            'table__zone'
        ).prefetch_related(
            Prefetch('orderitem_set', queryset=OrderItem.objects.select_related(
                'recipe__group'
            ).prefetch_related(
                'recipe__recipeitem_set__ingredient__unit'
            )),
            'payments',
            'container_sales__container'
        ).order_by('paid_at')
        
        return self._process_detailed_order_data(paid_orders, include_all_items=True)
    
    def _get_unsold_recipes_period(self, detailed_data, period_dates):
        """
        Obtiene las recetas que no fueron vendidas en el período seleccionado
        """
        from inventory.models import Recipe
        
        # Obtener IDs de recetas vendidas en este período
        sold_recipe_ids = set()
        for row in detailed_data:
            if row.get('recipe_name') and row.get('recipe_name') != 'Sin receta':
                try:
                    recipe = Recipe.objects.filter(name=row['recipe_name']).first()
                    if recipe:
                        sold_recipe_ids.add(recipe.id)
                except:
                    pass
        
        # Obtener todas las recetas activas que no fueron vendidas
        unsold_recipes = Recipe.objects.select_related('group').exclude(
            id__in=sold_recipe_ids
        ).filter(
            is_active=True
        ).order_by('name')
        
        unsold_list = []
        for recipe in unsold_recipes:
            unsold_list.append({
                'name': recipe.name,
                'category': recipe.group.name if recipe.group else 'Sin categoría',
                'price': float(recipe.base_price),
                'preparation_time': recipe.preparation_time or 0
            })
        
        return unsold_list
    
    def _get_detailed_order_data(self, selected_date):
        """
        Obtiene datos completamente detallados por fila (método legacy para compatibilidad)
        """
        # Usar el nuevo método con período 'all' para mantener compatibilidad
        period_dates = {
            'start_date': None,
            'end_date': None,
            'display_date': selected_date,
            'total_days': None
        }
        return self._get_detailed_order_data_period(period_dates)
    
    def _get_all_detailed_order_data(self, selected_date):
        """
        Obtiene TODOS los datos detallados para una fecha específica
        """
        # Obtener órdenes PAID para la fecha específica
        paid_orders = Order.objects.filter(
            status='PAID',
            paid_at__date=selected_date
        ).select_related(
            'table__zone'
        ).prefetch_related(
            Prefetch('orderitem_set', queryset=OrderItem.objects.select_related(
                'recipe__group'
            ).prefetch_related(
                'recipe__recipeitem_set__ingredient__unit'
            )),
            'payment_set'
        )
        
        # Procesar con include_all_items=True para incluir TODOS los estados
        return self._process_detailed_order_data(paid_orders, include_all_items=True)
    
    def _get_unsold_recipes(self, detailed_data, selected_date):
        """
        Obtiene las recetas no vendidas (método legacy para compatibilidad)
        """
        # Usar el nuevo método con período adecuado
        period_dates = {
            'start_date': None,
            'end_date': None,
            'display_date': selected_date,
            'total_days': None
        }
        return self._get_unsold_recipes_period(detailed_data, period_dates)
    
    def _process_detailed_order_data(self, paid_orders, include_all_items=False):
        """
        Procesa los datos de órdenes y devuelve filas detalladas
        """
        detailed_rows = []
        
        for order in paid_orders:
            # Información básica de la orden
            order_info = {
                'order_id': order.id,
                'table_number': order.table.table_number if order.table else 'N/A',
                'zone_name': order.table.zone.name if order.table and order.table.zone else 'N/A',
                'waiter': order.waiter or 'Sin asignar',
                'order_status': order.status,
                'order_total': float(order.total_amount),
                'created_at': order.created_at.isoformat() if order.created_at else None,
                'served_at': order.served_at.isoformat() if order.served_at else None,
                'paid_at': order.paid_at.isoformat() if order.paid_at else None,
                'service_time_minutes': None
            }
            
            # Calcular tiempo de servicio
            if order.created_at and order.paid_at:
                service_time_seconds = (order.paid_at - order.created_at).total_seconds()
                service_time_minutes = max(1, int(service_time_seconds / 60))
                order_info['service_time_minutes'] = service_time_minutes
            
            # Información de pagos
            payments_info = []
            for payment in order.payments.all():
                payments_info.append({
                    'payment_method': payment.payment_method,
                    'payment_amount': float(payment.amount),
                    'payment_date': payment.created_at.isoformat(),
                    'tax_amount': float(payment.tax_amount),
                    'payer_name': payment.payer_name or ''
                })
            
            # Información de envases
            containers_info = []
            for container_sale in order.container_sales.all():
                containers_info.append({
                    'container_name': container_sale.container.name,
                    'container_quantity': container_sale.quantity,
                    'container_unit_price': float(container_sale.unit_price),
                    'container_total_price': float(container_sale.total_price)
                })
            
            # Filtrar items según el modo
            if include_all_items:
                order_items = order.orderitem_set.all()
            else:
                order_items = [item for item in order.orderitem_set.all() if item.status == 'PAID']
            
            for order_item in order_items:
                # Información del item
                item_info = {
                    'item_id': order_item.id,
                    'recipe_name': order_item.recipe.name if order_item.recipe else 'Sin receta',
                    'recipe_version': order_item.recipe.version if order_item.recipe else 'N/A',
                    'category': order_item.recipe.group.name if order_item.recipe and order_item.recipe.group else 'Sin categoría',
                    'item_quantity': order_item.quantity,
                    'item_unit_price': float(order_item.unit_price),
                    'item_total_price': float(order_item.total_price),
                    'item_status': order_item.status,
                    'item_notes': order_item.notes or '',
                    'is_takeaway': order_item.is_takeaway,
                    'has_taper': order_item.has_taper,
                    'preparation_time': order_item.recipe.preparation_time if order_item.recipe else 0
                }
                
                # Ingredientes base de la receta
                recipe_ingredients = []
                if order_item.recipe:
                    for recipe_item in order_item.recipe.recipeitem_set.all():
                        ingredient_cost = float(recipe_item.ingredient.unit_price * recipe_item.quantity)
                        recipe_ingredients.append({
                            'ingredient_name': recipe_item.ingredient.name,
                            'ingredient_unit': recipe_item.ingredient.unit.name,
                            'ingredient_quantity': float(recipe_item.quantity),
                            'ingredient_unit_price': float(recipe_item.ingredient.unit_price),
                            'ingredient_total_cost': ingredient_cost,
                            'ingredient_type': 'base'
                        })
                
                # Ingredientes personalizados - funcionalidad removida
                custom_ingredients = []
                
                # Calcular costos
                total_ingredient_cost = sum(ing['ingredient_total_cost'] for ing in recipe_ingredients + custom_ingredients)
                profit_amount = float(order_item.total_price) - total_ingredient_cost
                profit_percentage = (profit_amount / total_ingredient_cost * 100) if total_ingredient_cost > 0 else 0
                
                # Crear fila detallada completa
                detailed_row = {
                    **order_info,
                    **item_info,
                    'total_ingredient_cost': total_ingredient_cost,
                    'profit_amount': profit_amount,
                    'profit_percentage': profit_percentage,
                    'recipe_ingredients': recipe_ingredients,
                    'custom_ingredients': custom_ingredients,
                    'all_ingredients': recipe_ingredients + custom_ingredients,
                    'payments': payments_info,
                    'containers': containers_info
                }
                
                detailed_rows.append(detailed_row)
        
        return detailed_rows
    
    def _calculate_item_status_breakdown(self, detailed_data):
        """
        Calcula la distribución de estados de items para la gráfica
        """
        status_counts = {}
        status_amounts = {}
        
        for row in detailed_data:
            status = row['item_status']
            amount = row['item_total_price']
            
            status_counts[status] = status_counts.get(status, 0) + 1
            status_amounts[status] = status_amounts.get(status, 0) + amount
        
        total_items = sum(status_counts.values())
        total_amount = sum(status_amounts.values())
        
        breakdown = []
        for status in ['CREATED', 'PREPARING', 'SERVED', 'PAID']:
            count = status_counts.get(status, 0)
            amount = status_amounts.get(status, 0)
            
            count_percentage = (count / total_items * 100) if total_items > 0 else 0
            amount_percentage = (amount / total_amount * 100) if total_amount > 0 else 0
            
            breakdown.append({
                'status': status,
                'count': count,
                'amount': amount,
                'count_percentage': count_percentage,
                'amount_percentage': amount_percentage
            })
        
        return breakdown
    
    def _get_unsold_recipes(self, detailed_data, selected_date):
        """
        Obtiene las recetas que no fueron vendidas en la fecha seleccionada
        """
        from inventory.models import Recipe
        
        # Obtener IDs de recetas vendidas en esta fecha
        sold_recipe_ids = set()
        for row in detailed_data:
            if row.get('recipe_name') and row.get('recipe_name') != 'Sin receta':
                # Buscar el ID de la receta por nombre
                try:
                    recipe = Recipe.objects.filter(name=row['recipe_name']).first()
                    if recipe:
                        sold_recipe_ids.add(recipe.id)
                except:
                    pass
        
        # Obtener todas las recetas activas que no fueron vendidas
        unsold_recipes = Recipe.objects.select_related('group').exclude(
            id__in=sold_recipe_ids
        ).filter(
            is_active=True  # Solo recetas activas
        ).order_by('name')  # Sin límite - mostrar todas las recetas no vendidas
        
        unsold_list = []
        for recipe in unsold_recipes:
            unsold_list.append({
                'name': recipe.name,
                'category': recipe.group.name if recipe.group else 'Sin categoría',
                'price': float(recipe.base_price),
                'preparation_time': recipe.preparation_time or 0
            })
        
        return unsold_list
    
    def _generate_dashboard_from_details(self, detailed_data, selected_date):
        """
        Genera el dashboard completo desde los datos detallados
        """
        if not detailed_data:
            return {
                'date': selected_date.isoformat(),
                'detailed_data': [],
                'summary': {'total_orders': 0, 'total_revenue': 0, 'average_ticket': 0, 'average_service_time': 0},
                'category_breakdown': [], 'top_dishes': [], 'waiter_performance': [],
                'zone_performance': [], 'top_tables': [], 'payment_methods': []
            }
        
        # Calcular métricas desde datos detallados
        unique_orders = {}
        total_revenue = Decimal('0')
        service_times = []
        category_stats = {}
        dish_stats = {}
        waiter_stats = {}
        zone_stats = {}
        table_stats = {}
        payment_method_stats = {}
        payment_method_counts = {}
        
        for row in detailed_data:
            order_id = row['order_id']
            
            # Stats por categoría (siempre)
            category = row['category']
            if category not in category_stats:
                category_stats[category] = {'revenue': Decimal('0'), 'quantity': 0}
            category_stats[category]['revenue'] += Decimal(str(row['item_total_price']))
            category_stats[category]['quantity'] += row['item_quantity']
            
            # Stats por plato (siempre)
            dish = row['recipe_name']
            if dish not in dish_stats:
                dish_stats[dish] = {
                    'category': category,
                    'quantity': 0,
                    'revenue': Decimal('0'),
                    'unit_price': Decimal(str(row['item_unit_price']))
                }
            dish_stats[dish]['quantity'] += row['item_quantity']
            dish_stats[dish]['revenue'] += Decimal(str(row['item_total_price']))
            
            # Stats por mesero y orden (solo una vez por orden)
            if order_id not in unique_orders:
                waiter = row['waiter']
                if waiter not in waiter_stats:
                    waiter_stats[waiter] = {'orders': 0, 'revenue': Decimal('0')}
                waiter_stats[waiter]['orders'] += 1
                waiter_stats[waiter]['revenue'] += Decimal(str(row['order_total']))
                
                # Agregar orden a unique_orders después de procesarla
                unique_orders[order_id] = {
                    'total': Decimal(str(row['order_total'])),
                    'waiter': row['waiter'],
                    'zone': row['zone_name'],
                    'table': row['table_number'],
                    'service_time': row['service_time_minutes']
                }
                total_revenue += Decimal(str(row['order_total']))
                
                if row['service_time_minutes']:
                    service_times.append(row['service_time_minutes'])
                
                # Stats de pagos (solo para la primera fila de cada orden) - OPTIMIZADO
                for payment in row['payments']:
                    method = payment['payment_method']
                    amount = Decimal(str(payment['payment_amount']))
                    payment_method_stats[method] = payment_method_stats.get(method, Decimal('0')) + amount
                    payment_method_counts[method] = payment_method_counts.get(method, 0) + 1
                
                # Stats por zona (solo una vez por orden)
                zone = row['zone_name']
                if zone not in zone_stats:
                    zone_stats[zone] = {'orders': 0, 'revenue': Decimal('0'), 'tables': set()}
                zone_stats[zone]['orders'] += 1
                zone_stats[zone]['revenue'] += Decimal(str(row['order_total']))
                zone_stats[zone]['tables'].add(row['table_number'])
                
                # Stats por mesa (solo una vez por orden)
                table = f"Mesa {row['table_number']}"
                if table not in table_stats:
                    table_stats[table] = Decimal('0')
                table_stats[table] += Decimal(str(row['order_total']))
        
        # Métricas generales
        total_orders = len(unique_orders)
        average_ticket = total_revenue / total_orders if total_orders > 0 else Decimal('0')
        if service_times:
            average_service_time = sum(service_times) / len(service_times)
        else:
            average_service_time = 0
        
        # Procesar stats (igual que antes pero desde datos detallados)
        # Categorías
        total_category_revenue = sum(cat['revenue'] for cat in category_stats.values())
        category_breakdown = []
        for category, stats in sorted(category_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            percentage = (stats['revenue'] / total_category_revenue * 100) if total_category_revenue > 0 else 0
            category_breakdown.append({
                'category': category,
                'revenue': float(stats['revenue']),
                'quantity': stats['quantity'],
                'percentage': float(percentage)
            })
        
        # Top platos
        top_dishes = []
        for dish, stats in sorted(dish_stats.items(), key=lambda x: x[1]['quantity'], reverse=True)[:10]:
            top_dishes.append({
                'name': dish,
                'category': stats['category'],
                'quantity': stats['quantity'],
                'revenue': float(stats['revenue']),
                'unit_price': float(stats['unit_price'])
            })
        
        # Performance meseros
        waiter_performance = []
        for waiter, stats in sorted(waiter_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)[:5]:
            avg_ticket = stats['revenue'] / stats['orders'] if stats['orders'] > 0 else Decimal('0')
            waiter_performance.append({
                'waiter': waiter,
                'orders': stats['orders'],
                'revenue': float(stats['revenue']),
                'average_ticket': float(avg_ticket)
            })
        
        # Performance por zonas
        zone_performance = []
        for zone, stats in sorted(zone_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            tables_used = len(stats['tables'])
            avg_per_table = stats['revenue'] / tables_used if tables_used > 0 else Decimal('0')
            zone_performance.append({
                'zone': zone,
                'orders': stats['orders'],
                'revenue': float(stats['revenue']),
                'tables_used': tables_used,
                'average_per_table': float(avg_per_table)
            })
        
        # Top 5 mesas
        top_tables = []
        for table, revenue in sorted(table_stats.items(), key=lambda x: x[1], reverse=True)[:5]:
            top_tables.append({
                'table': table,
                'revenue': float(revenue)
            })
        
        # Métodos de pago
        payment_methods = []
        for method, amount in payment_method_stats.items():
            percentage = (amount / total_revenue * 100) if total_revenue > 0 else 0
            payment_methods.append({
                'method': method,
                'amount': float(amount),
                'percentage': float(percentage),
                'transaction_count': payment_method_counts.get(method, 0)
            })
        
        
        return {
            'date': selected_date.isoformat(),
            'detailed_data': detailed_data,  # TODOS los datos detallados
            'summary': {
                'total_orders': total_orders,
                'total_revenue': float(total_revenue),
                'average_ticket': float(average_ticket),
                'average_service_time': float(average_service_time)
            },
            'category_breakdown': category_breakdown,
            'top_dishes': top_dishes,
            'waiter_performance': waiter_performance,
            'zone_performance': zone_performance,
            'top_tables': top_tables,
            'payment_methods': payment_methods
        }
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporta datos del dashboard a Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            from decimal import Decimal
            
            # Obtener fecha del parámetro
            date_param = request.query_params.get('date')
            if date_param:
                try:
                    selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = timezone.now().date()
            else:
                selected_date = timezone.now().date()
            
            # Obtener órdenes del día
            orders = Order.objects.filter(
                status='PAID',
                paid_at__date=selected_date
            ).select_related('table__zone').prefetch_related(
                Prefetch('orderitem_set', queryset=OrderItem.objects.select_related(
                    'recipe__group'
                ).prefetch_related('recipe__recipeitem_set__ingredient')),
                'payments'
            )
            
            # Crear workbook de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dashboard Operativo"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Fecha', 'Hora', 'ID Orden', 'Mesa', 'Zona', 'Mesero',
                'Receta', 'Categoría', 'Cantidad', 'Precio Unit.', 'Precio Total',
                'Estado', 'Notas', 'Delivery', 'Método Pago', 'Monto Pagado'
            ]
            
            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            row_num = 2
            total_revenue = Decimal('0')
            total_items = 0
            
            for order in orders:
                order_payments = list(order.payments.all())
                
                for item in order.orderitem_set.all():
                    # Calcular hora
                    hora = order.created_at.strftime('%H:%M') if order.created_at else ''
                    
                    # Datos del item
                    ws.cell(row=row_num, column=1, value=selected_date.strftime('%d/%m/%Y'))
                    ws.cell(row=row_num, column=2, value=hora)
                    ws.cell(row=row_num, column=3, value=order.id)
                    ws.cell(row=row_num, column=4, value=order.table.number if order.table else 'N/A')
                    ws.cell(row=row_num, column=5, value=order.table.zone.name if order.table and order.table.zone else 'N/A')
                    ws.cell(row=row_num, column=6, value=order.waiter or 'N/A')
                    ws.cell(row=row_num, column=7, value=item.recipe.name if item.recipe else 'N/A')
                    ws.cell(row=row_num, column=8, value=item.recipe.group.name if item.recipe and item.recipe.group else 'N/A')
                    ws.cell(row=row_num, column=9, value=item.quantity)
                    ws.cell(row=row_num, column=10, value=float(item.unit_price))
                    ws.cell(row=row_num, column=11, value=float(item.total_price))
                    ws.cell(row=row_num, column=12, value=item.status)
                    ws.cell(row=row_num, column=13, value=item.notes or '')
                    ws.cell(row=row_num, column=14, value='Sí' if item.is_takeaway else 'No')
                    
                    # Pagos (puede haber múltiples)
                    if order_payments:
                        payment = order_payments[0]
                        ws.cell(row=row_num, column=15, value=payment.payment_method)
                        ws.cell(row=row_num, column=16, value=float(payment.amount))
                    else:
                        ws.cell(row=row_num, column=15, value='N/A')
                        ws.cell(row=row_num, column=16, value=0)
                    
                    # Aplicar bordes
                    for col in range(1, 17):
                        ws.cell(row=row_num, column=col).border = border
                    
                    # Totales
                    if item.status == 'PAID':
                        total_revenue += item.total_price
                        total_items += item.quantity
                    
                    row_num += 1
            
            # Resumen al final
            row_num += 2
            ws.cell(row=row_num, column=1, value="RESUMEN").font = Font(bold=True)
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Items Vendidos:")
            ws.cell(row=row_num, column=2, value=total_items)
            row_num += 1
            ws.cell(row=row_num, column=1, value="Ingresos Totales:")
            ws.cell(row=row_num, column=2, value=float(total_revenue))
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Preparar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename=Dashboard_Operativo_{selected_date}.xlsx'
            
            # Guardar el workbook en la respuesta
            wb.save(response)
            
            return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error generating export: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_detailed_csv(self, response_data):
        """
        Genera CSV con datos completamente detallados
        """
        import csv
        from django.http import HttpResponse
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers completos
        headers = [
            'Fecha', 'ID_Orden', 'Mesa', 'Zona', 'Mesero', 'Estado_Orden', 'Total_Orden',
            'Creada_En', 'Servida_En', 'Pagada_En', 'Tiempo_Servicio_Min',
            'ID_Item', 'Receta', 'Version_Receta', 'Categoria', 'Cantidad_Item', 
            'Precio_Unitario', 'Precio_Total_Item', 'Estado_Item', 'Notas', 
            'Para_Llevar', 'Con_Envase', 'Tiempo_Preparacion',
            'Costo_Total_Ingredientes', 'Ganancia', 'Porcentaje_Ganancia',
            'Ingrediente', 'Tipo_Ingrediente', 'Unidad', 'Cantidad_Ingrediente', 
            'Precio_Unit_Ingrediente', 'Costo_Total_Ingrediente',
            'Metodo_Pago', 'Monto_Pago', 'Nombre_Pagador',
            'Envase_Nombre', 'Envase_Cantidad', 'Envase_Precio'
        ]
        
        writer.writerow(headers)
        
        # Escribir datos detallados
        for row in response_data.get('detailed_data', []):
            # Datos base por cada ingrediente
            all_ingredients = row.get('all_ingredients', [])
            payments = row.get('payments', [])
            containers = row.get('containers', [])
            
            if not all_ingredients:
                # Si no hay ingredientes, crear una fila con datos básicos
                base_row = [
                    response_data['date'], row['order_id'], row['table_number'], row['zone_name'],
                    row['waiter'], row['order_status'], row['order_total'],
                    row['created_at'], row['served_at'], row['paid_at'], row['service_time_minutes'],
                    row['item_id'], row['recipe_name'], row['recipe_version'], row['category'],
                    row['item_quantity'], row['item_unit_price'], row['item_total_price'],
                    row['item_status'], row['item_notes'], row['is_takeaway'], row['has_taper'],
                    row['preparation_time'], row['total_ingredient_cost'], row['profit_amount'],
                    row['profit_percentage'], '', '', '', '', '', '',
                    payments[0]['payment_method'] if payments else '',
                    payments[0]['payment_amount'] if payments else '',
                    payments[0]['payer_name'] if payments else '',
                    containers[0]['container_name'] if containers else '',
                    containers[0]['container_quantity'] if containers else '',
                    containers[0]['container_total_price'] if containers else ''
                ]
                writer.writerow(base_row)
            else:
                # Una fila por cada ingrediente - CORREGIDO: item_total_price solo en primera fila
                for idx, ingredient in enumerate(all_ingredients):
                    # Solo incluir item_total_price en la primera fila del item para evitar duplicación
                    item_total_display = row['item_total_price'] if idx == 0 else 0
                    
                    ingredient_row = [
                        response_data['date'], row['order_id'], row['table_number'], row['zone_name'],
                        row['waiter'], row['order_status'], row['order_total'],
                        row['created_at'], row['served_at'], row['paid_at'], row['service_time_minutes'],
                        row['item_id'], row['recipe_name'], row['recipe_version'], row['category'],
                        row['item_quantity'], row['item_unit_price'], item_total_display,
                        row['item_status'], row['item_notes'], row['is_takeaway'], row['has_taper'],
                        row['preparation_time'], row['total_ingredient_cost'], row['profit_amount'],
                        row['profit_percentage'],
                        ingredient['ingredient_name'], ingredient['ingredient_type'],
                        ingredient['ingredient_unit'], ingredient['ingredient_quantity'],
                        ingredient['ingredient_unit_price'], ingredient['ingredient_total_cost'],
                        payments[0]['payment_method'] if payments else '',
                        payments[0]['payment_amount'] if payments else '',
                        payments[0]['payer_name'] if payments else '',
                        containers[0]['container_name'] if containers else '',
                        containers[0]['container_quantity'] if containers else '',
                        containers[0]['container_total_price'] if containers else ''
                    ]
                    writer.writerow(ingredient_row)
        
        # Preparar respuesta
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"dashboard_detallado_{response_data['date']}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(output.getvalue())
        
        return response
    
    def _generate_detailed_excel(self, response_data):
        """
        Genera Excel con datos detallados en múltiples hojas
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        
        # Hoja 1: Datos Detallados Completos
        ws_details = wb.active
        ws_details.title = "Datos Detallados"
        
        # Headers (mismos que CSV)
        headers = [
            'Fecha', 'ID_Orden', 'Mesa', 'Zona', 'Mesero', 'Estado_Orden', 'Total_Orden',
            'Creada_En', 'Servida_En', 'Pagada_En', 'Tiempo_Servicio_Min',
            'ID_Item', 'Receta', 'Version_Receta', 'Categoria', 'Cantidad_Item', 
            'Precio_Unitario', 'Precio_Total_Item', 'Estado_Item', 'Notas', 
            'Para_Llevar', 'Con_Envase', 'Tiempo_Preparacion',
            'Costo_Total_Ingredientes', 'Ganancia', 'Porcentaje_Ganancia',
            'Ingrediente', 'Tipo_Ingrediente', 'Unidad', 'Cantidad_Ingrediente', 
            'Precio_Unit_Ingrediente', 'Costo_Total_Ingrediente'
        ]
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="366092")
        
        # Escribir datos
        row_num = 2
        for row in response_data.get('detailed_data', []):
            all_ingredients = row.get('all_ingredients', [])
            
            if not all_ingredients:
                # Fila sin ingredientes
                data_row = [
                    response_data['date'], row['order_id'], row['table_number'], row['zone_name'],
                    row['waiter'], row['order_status'], row['order_total'],
                    row['created_at'], row['served_at'], row['paid_at'], row['service_time_minutes'],
                    row['item_id'], row['recipe_name'], row['recipe_version'], row['category'],
                    row['item_quantity'], row['item_unit_price'], row['item_total_price'],
                    row['item_status'], row['item_notes'], row['is_takeaway'], row['has_taper'],
                    row['preparation_time'], row['total_ingredient_cost'], row['profit_amount'],
                    row['profit_percentage'], '', '', '', '', '', ''
                ]
                for col, value in enumerate(data_row, 1):
                    ws_details.cell(row=row_num, column=col, value=value)
                row_num += 1
            else:
                # Una fila por ingrediente - CORREGIDO: item_total_price solo en primera fila
                for idx, ingredient in enumerate(all_ingredients):
                    # Solo incluir item_total_price en la primera fila del item para evitar duplicación
                    item_total_display = row['item_total_price'] if idx == 0 else 0
                    
                    data_row = [
                        response_data['date'], row['order_id'], row['table_number'], row['zone_name'],
                        row['waiter'], row['order_status'], row['order_total'],
                        row['created_at'], row['served_at'], row['paid_at'], row['service_time_minutes'],
                        row['item_id'], row['recipe_name'], row['recipe_version'], row['category'],
                        row['item_quantity'], row['item_unit_price'], item_total_display,
                        row['item_status'], row['item_notes'], row['is_takeaway'], row['has_taper'],
                        row['preparation_time'], row['total_ingredient_cost'], row['profit_amount'],
                        row['profit_percentage'],
                        ingredient['ingredient_name'], ingredient['ingredient_type'],
                        ingredient['ingredient_unit'], ingredient['ingredient_quantity'],
                        ingredient['ingredient_unit_price'], ingredient['ingredient_total_cost']
                    ]
                    for col, value in enumerate(data_row, 1):
                        ws_details.cell(row=row_num, column=col, value=value)
                    row_num += 1
        
        # Hoja 2: Resumen Dashboard
        ws_summary = wb.create_sheet("Resumen Dashboard")
        
        summary_data = [
            ("Total de Órdenes", response_data['summary']['total_orders']),
            ("Ingresos Totales", f"S/ {response_data['summary']['total_revenue']:.2f}"),
            ("Ticket Promedio", f"S/ {response_data['summary']['average_ticket']:.2f}"),
            ("Tiempo Servicio Promedio", f"{response_data['summary']['average_service_time']:.1f} min")
        ]
        
        for idx, (metric, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=idx, column=1, value=metric)
            ws_summary.cell(row=idx, column=2, value=value)
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"dashboard_detallado_{response_data['date']}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response