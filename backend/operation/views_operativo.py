from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.authentication import SessionAuthentication
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, Prefetch
from django.db import connection
from datetime import datetime, timedelta
from decimal import Decimal
from .models import Order, OrderItem, Payment
from inventory.models import Recipe


class DashboardOperativoViewSet(viewsets.ViewSet):
    """
    Vista específica para Dashboard Operativo
    Usa EXCLUSIVAMENTE dashboard_operativo_view para máximo rendimiento y consistencia
    """
    # Usar autenticación AWS Cognito configurada en settings
    
    @action(detail=False, methods=['get'])
    def report(self, request):
        """
        Endpoint específico para dashboard operativo usando EXCLUSIVAMENTE dashboard_operativo_view
        Consulta unificada a la vista consolidada para máximo rendimiento
        """
        try:
            # Obtener parámetros del request
            query_params = getattr(request, 'query_params', request.GET)
            date_param = query_params.get('date')
            
            # Calcular fecha de referencia
            if date_param:
                try:
                    selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = timezone.now().date()
            else:
                selected_date = timezone.now().date()
            
            # Usar método completo de dashboard con vista
            operational_data = self._query_dashboard_view(selected_date)
            
            # Agregar información de la fecha
            operational_data['date'] = selected_date.isoformat()
            operational_data['timestamp'] = timezone.now().isoformat()
            
            return Response(operational_data)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error en dashboard operativo: {str(e)}")
            
            return Response({
                'error': str(e),
                'date': timezone.now().date().isoformat(),
                'timestamp': timezone.now().isoformat(),
                'summary': {
                    'active_orders': 0, 
                    'pending_items': 0, 
                    'preparing_items': 0, 
                    'served_items': 0,
                    'overdue_items': 0,
                    'total_revenue_today': 0
                },
                'kitchen_status': [],
                'zone_activity': [],
                'recent_orders': [],
                'hourly_activity': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _query_fallback_data(self, selected_date):
        """
        Método simplificado que usa solo consultas básicas sin vista compleja
        """
        try:
            # Consultas muy básicas y seguras
            total_orders = Order.objects.filter(
                created_at__date=selected_date,
                status='PAID'
            ).count()
            
            total_revenue = Order.objects.filter(
                created_at__date=selected_date,
                status='PAID'
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            # Stats básicas de items
            all_items = OrderItem.objects.filter(order__created_at__date=selected_date)
            pending_items = all_items.filter(status='CREATED').count()
            preparing_items = all_items.filter(status='PREPARING').count()
            served_items = all_items.filter(status='SERVED').count()
            
            return {
                'summary': {
                    'total_orders': total_orders,
                    'total_revenue': float(total_revenue),
                    'average_ticket': float(total_revenue / total_orders) if total_orders > 0 else 0,
                    'total_items': all_items.count(),
                    'average_service_time': 0,
                    'active_orders': Order.objects.filter(
                        created_at__date=selected_date
                    ).exclude(status='PAID').count(),
                    'pending_items': pending_items,
                    'preparing_items': preparing_items, 
                    'served_items': served_items,
                    'overdue_items': 0,
                    'total_revenue_today': float(total_revenue),
                    'delivery_orders': 0,
                    'restaurant_orders': total_orders,
                    'delivery_revenue': 0,
                    'restaurant_revenue': float(total_revenue),
                    'delivery_items': 0,
                    'restaurant_items': all_items.count()
                },
                'kitchen_status': [],
                'zone_activity': [],
                'recent_orders': [],
                'hourly_activity': [],
                'category_breakdown': [],
                'delivery_category_breakdown': [],
                'top_dishes': [],
                'waiter_performance': [],
                'payment_methods': [],
                'item_status_breakdown': [],
                'unsold_recipes': []
            }
        except Exception as e:
            # Si todo falla, devolver datos vacíos pero válidos
            import logging
            logging.getLogger(__name__).error(f"Fallback también falló: {e}")
            return {
                'summary': {
                    'total_orders': 0, 'total_revenue': 0, 'average_ticket': 0, 'total_items': 0,
                    'average_service_time': 0, 'active_orders': 0, 'pending_items': 0,
                    'preparing_items': 0, 'served_items': 0, 'overdue_items': 0,
                    'total_revenue_today': 0, 'delivery_orders': 0, 
                    'restaurant_orders': 0, 'delivery_revenue': 0, 'restaurant_revenue': 0,
                    'delivery_items': 0, 'restaurant_items': 0
                },
                'kitchen_status': [], 'zone_activity': [], 'recent_orders': [], 'hourly_activity': [],
                'category_breakdown': [], 'delivery_category_breakdown': [],
                'top_dishes': [], 'waiter_performance': [], 'payment_methods': [], 
                'item_status_breakdown': [], 'unsold_recipes': []
            }
    
    def _query_dashboard_view(self, selected_date):
        """
        Usa EXCLUSIVAMENTE dashboard_operativo_view para todos los datos del dashboard operativo
        Arquitectura consolidada: Una sola fuente de verdad para máximo rendimiento y consistencia
        """
        from django.db import connection
        from decimal import Decimal
        from collections import defaultdict
        
        # CONSULTA DIRECTA SIN PARÁMETROS PROBLEMÁTICOS
        from django.db import connection
        import logging
        
        # Deshabilitar temporalmente el logging SQL de Django para evitar string formatting issues
        django_db_logger = logging.getLogger('django.db.backends')
        original_level = django_db_logger.level
        django_db_logger.setLevel(logging.ERROR)
        
        cursor = connection.cursor()
        
        try:
            # Query con parámetros seguros usando ? placeholder
            date_str = selected_date.strftime('%Y-%m-%d')
            
            # Query completamente literal - sin parámetros para evitar string formatting issues
            query = f"""
                SELECT 
                    order_id, 
                    order_total, 
                    order_status, 
                    waiter, 
                    operational_date,
                    item_id, 
                    quantity, 
                    unit_price, 
                    total_price, 
                    total_with_container, 
                    item_status, 
                    is_takeaway,
                    recipe_name, 
                    category_name, 
                    category_id,
                    payment_method,
                    payment_amount,
                    created_at,
                    paid_at
                FROM dashboard_operativo_view
                WHERE operational_date = '{date_str}'
                ORDER BY order_id, item_id
            """
            
            cursor.execute(query)
            all_data = cursor.fetchall()
            
        except Exception as e:
            cursor.close()
            raise Exception(f"Error en consulta dashboard: {str(e)}")
        finally:
            cursor.close()
            # Restaurar el logging SQL de Django
            django_db_logger.setLevel(original_level)
        
        if not all_data:
            # Sin datos para la fecha
            return {
                'summary': {
                    'total_orders': 0, 'total_revenue': 0, 'average_ticket': 0, 'total_items': 0,
                    'average_service_time': 0, 'active_orders': 0, 'pending_items': 0,
                    'preparing_items': 0, 'served_items': 0,
                    'delivery_orders': 0, 'restaurant_orders': 0,
                    'delivery_revenue': 0, 'restaurant_revenue': 0,
                    'delivery_items': 0, 'restaurant_items': 0
                },
                'category_breakdown': [], 
                'delivery_category_breakdown': [],
                'top_dishes': [], 
                'waiter_performance': [],
                'payment_methods': [], 
                'item_status_breakdown': [], 
                'unsold_recipes': []
            }
        
        # Inicializar estructuras de datos
        paid_orders = set()
        paid_orders_totals = {}
        category_stats = defaultdict(lambda: {'revenue': Decimal('0'), 'quantity': 0})
        delivery_category_stats = defaultdict(lambda: {'revenue': Decimal('0'), 'quantity': 0, 'recipes': {}})
        dish_stats = defaultdict(lambda: {'category': '', 'quantity': 0, 'revenue': Decimal('0'), 'unit_price': Decimal('0')})
        waiter_stats = defaultdict(lambda: {'revenue': Decimal('0'), 'orders': 0})
        item_status_stats = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})
        payment_stats = defaultdict(lambda: {'amount': Decimal('0'), 'count': 0})
        
        # Estadísticas de delivery/takeaway
        delivery_stats = {
            'total_delivery_orders': 0,
            'total_delivery_revenue': Decimal('0'),
            'total_restaurant_orders': 0, 
            'total_restaurant_revenue': Decimal('0'),
            'delivery_items': 0,
            'restaurant_items': 0
        }
        
        # Métricas operativas
        active_orders = set()
        pending_items = 0
        preparing_items = 0
        served_items = 0
        service_times = []
        
        # PROCESAR TODOS LOS DATOS DE LA VISTA CONSOLIDADA
        for row in all_data:
            # Desempaquetar con manejo de errores - exactamente 19 campos del SELECT
            try:
                (order_id, order_total, order_status, waiter, operational_date,
                 item_id, quantity, unit_price, total_price, total_with_container, item_status, is_takeaway,
                 recipe_name, category_name, category_id,
                 payment_method, payment_amount, created_at, paid_at) = row
            except ValueError as e:
                # Debug: imprimir la estructura de datos recibida
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error desempaquetando fila de datos: {e}. Fila: {row}")
                continue
            
            # Procesar solo órdenes PAID para métricas financieras
            if order_status == 'PAID' and item_id:
                paid_orders.add(order_id)
                
                # Guardar total de la orden una sola vez
                if order_id not in paid_orders_totals:
                    paid_orders_totals[order_id] = Decimal(str(order_total or 0))
                
                # Stats por categoría (solo órdenes PAID) - USAR total_with_container para consistencia
                if category_name and recipe_name:
                    category = category_name or 'Sin Categoría'
                    category_stats[category]['revenue'] += Decimal(str(total_with_container or 0))
                    category_stats[category]['quantity'] += quantity or 0
                
                # Stats por receta (solo órdenes PAID) - USAR total_with_container para consistencia
                if recipe_name:
                    dish_stats[recipe_name]['category'] = category_name or 'Sin Categoría'
                    dish_stats[recipe_name]['quantity'] += quantity or 0
                    dish_stats[recipe_name]['revenue'] += Decimal(str(total_with_container or 0))
                    dish_stats[recipe_name]['unit_price'] = Decimal(str(unit_price or 0))
                
                # Stats por delivery/takeaway (solo órdenes PAID)
                if is_takeaway:
                    delivery_stats['delivery_items'] += quantity or 0
                    
                    # Stats específicos de delivery por categoría - USAR total_with_container para consistencia
                    if category_name and recipe_name:
                        category = category_name or 'Sin Categoría'
                        delivery_category_stats[category]['revenue'] += Decimal(str(total_with_container or 0))
                        delivery_category_stats[category]['quantity'] += quantity or 0
                        
                        # Agregar recipe a la categoría delivery
                        if recipe_name not in delivery_category_stats[category]['recipes']:
                            delivery_category_stats[category]['recipes'][recipe_name] = {
                                'quantity': 0,
                                'revenue': Decimal('0'),
                                'unit_price': Decimal(str(unit_price or 0))
                            }
                        
                        delivery_category_stats[category]['recipes'][recipe_name]['quantity'] += quantity or 0
                        delivery_category_stats[category]['recipes'][recipe_name]['revenue'] += Decimal(str(total_with_container or 0))
                else:
                    delivery_stats['restaurant_items'] += quantity or 0
                
                # Stats por mesero (solo órdenes PAID) - ACUMULAR por items para consistencia con total_with_container
                waiter_name = waiter or 'Sin Asignar'
                waiter_stats[waiter_name]['revenue'] += Decimal(str(total_with_container or 0))
                
                # Contar órdenes únicas por mesero
                if order_id not in [w['order_id'] for w in waiter_stats[waiter_name].get('processed_orders', [])]:
                    if 'processed_orders' not in waiter_stats[waiter_name]:
                        waiter_stats[waiter_name]['processed_orders'] = []
                    waiter_stats[waiter_name]['processed_orders'].append({'order_id': order_id})
                    waiter_stats[waiter_name]['orders'] += 1
            
            # Calcular delivery/restaurant por item individual - USAR total_with_container para consistencia
            if order_status == 'PAID' and item_id:
                if is_takeaway:
                    delivery_stats['total_delivery_revenue'] += Decimal(str(total_with_container or 0))
                else:
                    delivery_stats['total_restaurant_revenue'] += Decimal(str(total_with_container or 0))
            
            # Estado de items: TODOS los items del día (no solo PAID) - USAR total_with_container para consistencia
            if item_id and item_status:
                item_status_stats[item_status]['count'] += 1
                if order_status == 'PAID':
                    item_status_stats[item_status]['amount'] += Decimal(str(total_with_container or 0))
            
            # Métricas operativas: todas las órdenes activas
            if order_status in ('CREATED', 'SERVED'):
                active_orders.add(order_id)
            
            if item_status == 'CREATED':
                pending_items += 1
            elif item_status == 'PREPARING':
                preparing_items += 1
            elif item_status == 'SERVED':
                served_items += 1
            
            # Stats de pagos (solo órdenes PAID) - USAR payment_method y payment_amount para análisis simplificado
            if payment_method and payment_amount and order_status == 'PAID' and order_id not in [order_id for stat in payment_stats.values() for order_id in stat.get('processed_orders', [])]:
                # El payment_method viene directamente de la tabla payment
                if payment_method and payment_method != 'Sin pagos':
                    try:
                        amount = Decimal(str(payment_amount or 0))
                        if amount > 0:
                            payment_stats[payment_method]['amount'] += amount
                            payment_stats[payment_method]['count'] += 1
                            if 'processed_orders' not in payment_stats[payment_method]:
                                payment_stats[payment_method]['processed_orders'] = []
                            payment_stats[payment_method]['processed_orders'].append(order_id)
                    except (ValueError, TypeError):
                        continue
                
                # Si no hay payment_method pero sí payment_amount, usar método genérico
                elif payment_amount and payment_amount > 0:
                    payment_stats['UNKNOWN']['amount'] += Decimal(str(payment_amount))
                    payment_stats['UNKNOWN']['count'] += 1
                    if 'processed_orders' not in payment_stats['UNKNOWN']:
                        payment_stats['UNKNOWN']['processed_orders'] = []
                    payment_stats['UNKNOWN']['processed_orders'].append(order_id)
            
            # Calcular tiempo de servicio para órdenes completadas
            if order_status == 'PAID' and order_id not in [entry['order_id'] for entry in service_times]:
                # Usar datos directos de la fila actual
                if created_at and paid_at:
                    # Calcular diferencia en minutos - USAR TIEMPO REAL
                    time_diff = (paid_at - created_at).total_seconds() / 60
                    
                    # Usar el tiempo REAL siempre (sin estimaciones)
                    # Solo filtrar tiempos completamente irreales (> 7 días)
                    if 0 < time_diff < 10080:  # Entre 0 minutos y 7 días
                        service_times.append({
                            'order_id': order_id,
                            'service_time': time_diff
                        })
        
        # Contar órdenes únicas por tipo
        order_has_delivery_items = set()
        order_has_local_items = set()
        
        for row in all_data:
            # Use the same unpacking as the main loop (19 fields) - INCLUIR created_at, paid_at
            (order_id, order_total, order_status, waiter, operational_date,
             item_id, quantity, unit_price, total_price, total_with_container, item_status, is_takeaway,
             recipe_name, category_name, category_id,
             payment_method, payment_amount, created_at, paid_at) = row
            
            if order_status == 'PAID' and item_id:
                if is_takeaway:
                    order_has_delivery_items.add(order_id)
                else:
                    order_has_local_items.add(order_id)
        
        delivery_stats['total_delivery_orders'] = len(order_has_delivery_items)
        delivery_stats['total_restaurant_orders'] = len(order_has_local_items)
        
        # Calcular métricas principales
        total_orders = len(paid_orders)
        # USAR la suma de delivery + restaurant revenue para consistencia con total_with_container
        total_revenue = delivery_stats['total_delivery_revenue'] + delivery_stats['total_restaurant_revenue']
        average_ticket = total_revenue / total_orders if total_orders > 0 else Decimal('0')
        # Contar items de órdenes PAID: row[2] = order_status, row[5] = item_id
        total_items = len([row for row in all_data if row[2] == 'PAID' and row[5] is not None])
        average_service_time = sum(entry['service_time'] for entry in service_times) / len(service_times) if service_times else 0
        
        # Category breakdown
        total_category_revenue = sum(cat['revenue'] for cat in category_stats.values())
        category_breakdown = []
        for category, stats in sorted(category_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            percentage = (stats['revenue'] / total_category_revenue * 100) if total_category_revenue > 0 else 0
            category_breakdown.append({
                'category': category,
                'revenue': float(stats['revenue']),
                'quantity': stats['quantity'],
                'percentage': float(percentage)
            })
        
        # Delivery Category breakdown
        total_delivery_revenue = sum(cat['revenue'] for cat in delivery_category_stats.values())
        delivery_category_breakdown = []
        for category, stats in sorted(delivery_category_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            percentage = (stats['revenue'] / total_delivery_revenue * 100) if total_delivery_revenue > 0 else 0
            
            # Formatear recipes para el frontend
            recipes_list = []
            for recipe_name, recipe_stats in sorted(stats['recipes'].items(), key=lambda x: x[1]['revenue'], reverse=True):
                recipes_list.append({
                    'name': recipe_name,
                    'quantity': recipe_stats['quantity'],
                    'revenue': float(recipe_stats['revenue']),
                    'unit_price': float(recipe_stats['unit_price'])
                })
            
            delivery_category_breakdown.append({
                'category': category,
                'revenue': float(stats['revenue']),
                'quantity': stats['quantity'],
                'percentage': float(percentage),
                'recipes': recipes_list
            })
        
        # Top dishes
        top_dishes = []
        for dish, stats in sorted(dish_stats.items(), key=lambda x: x[1]['quantity'], reverse=True)[:10]:
            top_dishes.append({
                'name': dish,
                'category': stats['category'],
                'quantity': stats['quantity'],
                'revenue': float(stats['revenue']),
                'unit_price': float(stats['unit_price'])
            })
        
        # Waiter performance
        waiter_performance = []
        for waiter, stats in sorted(waiter_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            clean_stats = {k: v for k, v in stats.items() if k != 'processed_orders'}
            avg_ticket = clean_stats['revenue'] / clean_stats['orders'] if clean_stats['orders'] > 0 else Decimal('0')
            waiter_performance.append({
                'waiter': waiter,
                'revenue': float(clean_stats['revenue']),
                'orders': clean_stats['orders'],
                'average_ticket': float(avg_ticket)
            })
        
        # Payment methods
        total_payment_amount = sum(stat['amount'] for stat in payment_stats.values())
        payment_methods = []
        for method, stats in payment_stats.items():
            percentage = (stats['amount'] / total_payment_amount * 100) if total_payment_amount > 0 else 0
            # Limpiar stats para el frontend (remover processed_orders)
            clean_stats = {k: v for k, v in stats.items() if k != 'processed_orders'}
            payment_methods.append({
                'method': method,
                'amount': float(clean_stats['amount']),
                'percentage': float(percentage),
                'transaction_count': clean_stats['count']
            })
        
        # Item status breakdown
        total_items_status = sum(stat['count'] for stat in item_status_stats.values())
        item_status_breakdown = []
        for status_name, stats in item_status_stats.items():
            percentage = (stats['count'] / total_items_status * 100) if total_items_status > 0 else 0
            item_status_breakdown.append({
                'status': status_name,
                'count': stats['count'],
                'amount': float(stats['amount']),
                'count_percentage': float(percentage)
            })
        
        # Unsold recipes (obtener todas las recetas activas y compararlas)
        unsold_recipes_list = []
        try:
            all_recipes = Recipe.objects.filter(is_active=True, is_available=True).select_related('group')
            sold_recipe_names = set(dish_stats.keys())
            
            for recipe in all_recipes:
                if recipe.name not in sold_recipe_names:
                    unsold_recipes_list.append({
                        'name': recipe.name,
                        'category': recipe.group.name if recipe.group else 'Sin Categoría',
                        'price': float(recipe.base_price)
                    })
        except Exception as e:
            # Si hay error, continuar sin recetas no vendidas
            pass
        
        return {
            'summary': {
                'total_orders': total_orders,
                'total_revenue': float(total_revenue),
                'average_ticket': float(average_ticket),
                'total_items': total_items,
                'average_service_time': float(average_service_time),
                'active_orders': len(active_orders),
                'pending_items': pending_items,
                'preparing_items': preparing_items,
                'served_items': served_items,
                # Delivery/Restaurant breakdown
                'delivery_orders': delivery_stats['total_delivery_orders'],
                'restaurant_orders': delivery_stats['total_restaurant_orders'],
                'delivery_revenue': float(delivery_stats['total_delivery_revenue']),
                'restaurant_revenue': float(delivery_stats['total_restaurant_revenue']),
                'delivery_items': delivery_stats['delivery_items'],
                'restaurant_items': delivery_stats['restaurant_items']
            },
            'category_breakdown': category_breakdown,
            'delivery_category_breakdown': delivery_category_breakdown,
            'top_dishes': top_dishes,
            'waiter_performance': waiter_performance,
            'payment_methods': payment_methods,
            'item_status_breakdown': item_status_breakdown,
            'unsold_recipes': unsold_recipes_list
        }
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporta datos del dashboard operativo a Excel usando EXCLUSIVAMENTE dashboard_operativo_view
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            from django.db import connection
            from decimal import Decimal
            
            # Obtener fecha del parámetro
            date_param = request.query_params.get('date')
            if date_param:
                try:
                    selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = timezone.now().date()
            else:
                selected_date = timezone.now().date()
            
            # Obtener datos usando la vista consolidada
            dashboard_data = self._query_dashboard_view(selected_date)
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Dashboard Operativo {selected_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Solo tabla de datos - usando vista simplificada pero completa
            cursor = connection.cursor()
            cursor.execute("""
                SELECT 
                    DATE(o.created_at) as operational_date,
                    o.id as order_id,
                    o.status as order_status,
                    o.waiter,
                    o.table_number,
                    z.name as zone_name,
                    oi.id as item_id,
                    r.name as recipe_name,
                    g.name as category_name,
                    oi.quantity,
                    oi.unit_price,
                    oi.total_price,
                    (oi.total_price + COALESCE(cs.total_container_price, 0)) as total_with_container,
                    oi.status as item_status,
                    oi.is_takeaway,
                    p.payment_method,
                    p.amount as payment_amount,
                    0 as recipe_total_ingredient_cost,
                    0 as recipe_profit_margin,
                    0 as preparation_time,
                    0 as service_time_minutes,
                    'N/A' as meal_period,
                    'N/A' as day_of_week,
                    c.name as container_name,
                    cs.unit_price as container_unit_price,
                    -- Datos específicos del ingrediente (UNA FILA POR INGREDIENTE)
                    COALESCE(ing.name, 'Sin ingredientes definidos') as ingredient_name,
                    COALESCE(ri.quantity, 0) as ingredient_quantity,
                    COALESCE(ing.unit_price, 0) as ingredient_unit_price,
                    COALESCE(ri.quantity * ing.unit_price, 0) as ingredient_total_cost,
                    COALESCE(ri.quantity * ing.unit_price * oi.quantity, 0) as ingredient_cost_for_item,
                    COALESCE(ing.current_stock, 0) as ingredient_stock,
                    o.created_at,
                    p.paid_at
                FROM "order" o
                LEFT JOIN order_item oi ON o.id = oi.order_id
                LEFT JOIN recipe r ON oi.recipe_id = r.id
                LEFT JOIN "group" g ON r.group_id = g.id
                LEFT JOIN "table" t ON o.table_id = t.id
                LEFT JOIN zone z ON t.zone_id = z.id
                LEFT JOIN payment p ON o.id = p.order_id
                LEFT JOIN container_sale cs ON o.id = cs.order_id
                LEFT JOIN container c ON cs.container_id = c.id
                LEFT JOIN recipe_item ri ON r.id = ri.recipe_id
                LEFT JOIN ingredient ing ON ri.ingredient_id = ing.id
                WHERE DATE(o.created_at) = ?
                ORDER BY o.id, oi.id, ing.name
            """, [selected_date])
            
            raw_data = cursor.fetchall()
            cursor.close()
            
            # Headers expandidos - UNA FILA POR INGREDIENTE
            headers = [
                'Fecha', 'Order ID', 'Estado Orden', 'Mesero', 'Mesa', 'Zona',
                'Item ID', 'Receta', 'Categoría', 'Cantidad Item', 'Precio Unit. Item',
                'Total Item', 'Total c/Cont.', 'Estado Item', 'Es Delivery',
                'Método Pago', 'Monto Pago', 'Costo Total Ingredientes', 'Margen Ganancia',
                'Tiempo Prep (min)', 'Tiempo Servicio (min)', 'Período Comida', 'Día Semana',
                'Contenedor', 'Precio Contenedor', 
                'Ingrediente', 'Cantidad Ingrediente', 'Precio Unit. Ingrediente', 
                'Costo Total Ingrediente', 'Costo para este Item', 'Stock Ingrediente',
                'Creado', 'Pagado'
            ]
            
            # Headers de columnas (fila 1)
            row = 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos de la tabla (empezar en fila 2)
            row = 2
            for data_row in raw_data:
                (operational_date, order_id, order_status, waiter, table_number, zone_name,
                 item_id, recipe_name, category_name, quantity, unit_price, 
                 total_price, total_with_container, item_status, is_takeaway,
                 payment_method, payment_amount, recipe_total_ingredient_cost, recipe_profit_margin,
                 preparation_time, service_time_minutes, meal_period, day_of_week,
                 container_name, container_unit_price,
                 ingredient_name, ingredient_quantity, ingredient_unit_price, 
                 ingredient_total_cost, ingredient_cost_for_item, ingredient_stock,
                 created_at, paid_at) = data_row
                
                ws.cell(row=row, column=1).value = str(operational_date)
                ws.cell(row=row, column=2).value = order_id
                ws.cell(row=row, column=3).value = order_status
                ws.cell(row=row, column=4).value = waiter or 'Sin Asignar'
                ws.cell(row=row, column=5).value = table_number or 'N/A'
                ws.cell(row=row, column=6).value = zone_name or 'N/A'
                ws.cell(row=row, column=7).value = item_id
                ws.cell(row=row, column=8).value = recipe_name
                ws.cell(row=row, column=9).value = category_name
                ws.cell(row=row, column=10).value = quantity
                ws.cell(row=row, column=11).value = float(unit_price) if unit_price else 0
                ws.cell(row=row, column=12).value = float(total_price) if total_price else 0
                ws.cell(row=row, column=13).value = float(total_with_container) if total_with_container else 0
                ws.cell(row=row, column=14).value = item_status
                ws.cell(row=row, column=15).value = 'Sí' if is_takeaway else 'No'
                ws.cell(row=row, column=16).value = payment_method
                ws.cell(row=row, column=17).value = float(payment_amount) if payment_amount else 0
                ws.cell(row=row, column=18).value = float(recipe_total_ingredient_cost) if recipe_total_ingredient_cost else 0
                ws.cell(row=row, column=19).value = float(recipe_profit_margin) if recipe_profit_margin else 0
                ws.cell(row=row, column=20).value = preparation_time if preparation_time else 0
                ws.cell(row=row, column=21).value = service_time_minutes if service_time_minutes else 0
                ws.cell(row=row, column=22).value = meal_period
                ws.cell(row=row, column=23).value = day_of_week
                ws.cell(row=row, column=24).value = container_name or 'Sin contenedor'
                ws.cell(row=row, column=25).value = float(container_unit_price) if container_unit_price else 0
                # Datos específicos del ingrediente (una fila por ingrediente)
                ws.cell(row=row, column=26).value = ingredient_name or 'Sin ingredientes'
                ws.cell(row=row, column=27).value = float(ingredient_quantity) if ingredient_quantity else 0
                ws.cell(row=row, column=28).value = float(ingredient_unit_price) if ingredient_unit_price else 0
                ws.cell(row=row, column=29).value = float(ingredient_total_cost) if ingredient_total_cost else 0
                ws.cell(row=row, column=30).value = float(ingredient_cost_for_item) if ingredient_cost_for_item else 0
                ws.cell(row=row, column=31).value = float(ingredient_stock) if ingredient_stock else 0
                ws.cell(row=row, column=32).value = str(created_at) if created_at else ''
                ws.cell(row=row, column=33).value = str(paid_at) if paid_at else ''
                
                row += 1
            
            # Ajustar columnas para trazabilidad por ingrediente (33 columnas)
            column_widths = [12, 8, 12, 15, 8, 12, 8, 25, 15, 8, 10, 10, 10, 12, 10, 12, 10, 12, 12, 8, 8, 12, 12, 15, 10, 25, 10, 10, 12, 12, 10, 16, 16]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Crear respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="dashboard_operativo_{selected_date}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({'error': f'Error al generar Excel: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)